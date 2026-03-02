# -*- coding: utf-8 -*-
"""
PASSO 0 (opcional): Exporta o "mapeamento de campos por fase" da planilha Excel
para um JSON usado pela Supabase Edge Function.

Fonte: ../Mapeamento de eventos necessários - fase x campos por fase.xlsx
Aba esperada: uma que tenha cabeçalho "PIPE" e "ETAPA" nas colunas A e B.

Saída: integration/edge_function/required_fields_by_phase.json

Observações:
- O JSON guarda *labels normalizados* (sem acentos, minúsculo) dos campos
  conforme a planilha. A Edge Function converte esses labels em colunas da
  tabela `eventos` via `FIELD_LABEL_MAP`.
- A célula "Data e Horário da alteração/criação de fase" é ignorada (já vira `created_at`).
"""

import io
import json
import re
import sys
from pathlib import Path

import openpyxl


# Força UTF-8 no stdout para evitar erros no Windows (cp1252)
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


BASE = Path(__file__).parent
XLSX = BASE.parent / "Mapeamento de eventos necessários - fase x campos por fase.xlsx"
OUT  = BASE / "edge_function" / "required_fields_by_phase.json"


def normalize(text: str) -> str:
    import unicodedata

    s = str(text or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[?!,.]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_pipe(text: str) -> str:
    """
    A planilha às vezes usa prefixo "PIPE ...", mas o Pipefy geralmente retorna
    apenas o nome real do pipe (ex.: "SDR - COMERCIAL"). Para não dar mismatch,
    removemos um possível prefixo "pipe ".
    """
    s = normalize(text)
    if s.startswith("pipe "):
        s = s[len("pipe ") :].strip()
    return s


def find_mapping_sheet(wb: openpyxl.Workbook):
    for name in wb.sheetnames:
        ws = wb[name]
        a = normalize(ws.cell(1, 1).value)
        b = normalize(ws.cell(1, 2).value)
        if a == "pipe" and b == "etapa":
            return ws
    return None


def main() -> int:
    if not XLSX.exists():
        print(f"[!] Planilha não encontrada. Pulando export: {XLSX}")
        # Não falha o pipeline (passo opcional)
        return 0

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = find_mapping_sheet(wb)
    if ws is None:
        print("[!] Não encontrei uma aba com cabeçalho PIPE/ETAPA em A1/B1.")
        print(f"    Abas disponíveis: {wb.sheetnames}")
        return 1

    mapping: dict[str, dict[str, list[str]]] = {}

    for r in range(2, ws.max_row + 1):
        pipe = ws.cell(r, 1).value
        etapa = ws.cell(r, 2).value
        if not pipe or not etapa:
            continue

        pipe_key = normalize_pipe(pipe)
        etapa_key = normalize(etapa)

        labels: list[str] = []
        for c in range(3, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            v = str(v).strip()
            if not v:
                continue
            if "data e horario" in normalize(v):
                continue
            labels.append(normalize(v))

        # de-dup preservando ordem
        seen: set[str] = set()
        uniq: list[str] = []
        for l in labels:
            if l not in seen:
                seen.add(l)
                uniq.append(l)

        mapping.setdefault(pipe_key, {})[etapa_key] = uniq

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(mapping, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8")

    print(f"[OK] JSON exportado: {OUT}")
    print(f"     Pipes: {len(mapping)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

