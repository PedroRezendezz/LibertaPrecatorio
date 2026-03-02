# -*- coding: utf-8 -*-
"""
Gera o SQL da tabela public.mapeamento_campos_fase a partir da planilha
`Mapeamento de Eventos Necessários.xlsx` (aba `campos`).

Saída: sobrescreve `integration/sql/003_mapeamento_campos_fase.sql`
com:
- DROP das views/tabelas antigas pedidas pelo usuário
- CREATE TABLE mapeamento_campos_fase (pipe, fase, campo_label)
- INSERTs com todos os (pipe, fase, campo) da planilha
"""

import io
import re
import sys
from pathlib import Path

import openpyxl


if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Mapeamento de Eventos Necessários.xlsx"
OUT_SQL = ROOT / "integration" / "sql" / "003_mapeamento_campos_fase.sql"


def main() -> int:
    if not XLSX.exists():
        print(f"[!] Planilha não encontrada: {XLSX}")
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if "campos" not in wb.sheetnames:
        print(f"[!] Aba 'campos' não encontrada. Abas disponíveis: {wb.sheetnames}")
        return 1

    ws = wb["campos"]
    max_r = ws.max_row

    # Coleta triplets distintos (pipe, fase, campo_label)
    entries: set[tuple[str, str, str]] = set()
    for r in range(2, max_r + 1):
        pipe = ws.cell(r, 1).value
        fase = ws.cell(r, 2).value
        if not pipe or not fase:
            continue
        pipe_s = str(pipe).strip()
        fase_s = str(fase).strip()
        if not pipe_s or not fase_s:
            continue

        # Na nova planilha, a coluna C ("CAMPOS") contém um texto
        # com vários campos separados por ';' e/ou quebras de linha.
        raw_val = ws.cell(r, 3).value
        if raw_val is None:
            continue
        blob = str(raw_val)
        # Quebra por ';' e por quebras de linha
        parts = re.split(r"[;\n]+", blob)
        for part in parts:
            campo = part.strip()
            if not campo:
                continue
            if "data e horario" in campo.lower():
                continue
            entries.add((pipe_s, fase_s, campo))

    rows = sorted(entries, key=lambda t: (t[0].lower(), t[1].lower(), t[2].lower()))

    lines: list[str] = []
    lines.append(
        "-- Gerado automaticamente a partir de 'Mapeamento de Eventos Necessários.xlsx' (aba 'campos')"
    )
    lines.append("-- NÃO editar manualmente; rode generate_mapeamento_campos_sql.py se a planilha mudar.")
    lines.append("")
    # Drops pedidos pelo usuário
    lines.append("drop view if exists public.v_funil_sdr_comercial;")
    lines.append("drop view if exists public.v_mapeamento_campos_fase;")
    lines.append("drop table if exists public.mapeamento_campos_fase cascade;")
    lines.append("")

    lines.append("create table if not exists public.mapeamento_campos_fase (")
    lines.append("    id          bigserial primary key,")
    lines.append("    pipe        text not null,")
    lines.append("    fase        text not null,")
    lines.append("    campo_label text not null,")
    lines.append("    unique (pipe, fase, campo_label)")
    lines.append(");")
    lines.append("")

    lines.append("truncate table public.mapeamento_campos_fase;")
    lines.append("")

    if rows:
        lines.append("insert into public.mapeamento_campos_fase (pipe, fase, campo_label) values")
        value_lines = []
        for pipe_s, fase_s, label in rows:
            p = pipe_s.replace("'", "''")
            f = fase_s.replace("'", "''")
            l = label.replace("'", "''")
            value_lines.append(f"    ('{p}', '{f}', '{l}')")
        lines.append(",\n".join(value_lines) + ";")
        lines.append("")

    sql = "\n".join(lines) + "\n"
    OUT_SQL.write_text(sql, encoding="utf-8")
    print(f"[OK] SQL gerado em: {OUT_SQL}")
    print(f"     Registros: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

